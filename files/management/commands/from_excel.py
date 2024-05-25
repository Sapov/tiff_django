import openpyxl


def load_excel(sheet: str, range_cell: str):
    '''Читаем ексель файл в список списков
    передаем имя листа и диапазон таблицы'''
    # wb = openpyxl.load_workbook('price.xlsx')
    wb = openpyxl.load_workbook('files/management/commands/price.xlsx')
    # print(wb.sheetnames) # Посмотреть имена листов в Эксель файле
    sheet = wb[sheet]  # читаю первый лист по Ширке
    Material_data = []
    for row in sheet[range_cell]:
        item = [cell.value for cell in row]
        Material_data.append(item)
    # print(Material_data)
    return Material_data


if __name__ == '__main__':
    load_excel('shirka', 'b2:f9')
